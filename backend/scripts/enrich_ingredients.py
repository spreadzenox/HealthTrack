#!/usr/bin/env python3
"""
Enrichit backend/data/ingredients_nutrition.json à partir de bases ouvertes.

Sources supportées:
- Ciqual (ANSES, France): CSV téléchargeable sur recherche.data.gouv.fr
  https://entrepot.recherche.data.gouv.fr/dataset.xhtml?persistentId=doi:10.57745/RDMHWY
  Télécharger le fichier "Table Ciqual" en CSV ou Excel, puis extraire en CSV UTF-8.

Usage:
  python -m scripts.enrich_ingredients ciqual path/to/Table_Ciqual.csv
  python -m scripts.enrich_ingredients ciqual path/to/ciqual.csv --export-names 500
  python -m scripts.enrich_ingredients merge path/to/usda_export.json  # format HealthTrack
"""
from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Répertoire backend (parent de scripts/)
BACKEND_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BACKEND_DIR / "data"
INGREDIENTS_FILE = DATA_DIR / "ingredients_nutrition.json"
APP_NAMES_FILE = BACKEND_DIR.parent / "app" / "src" / "data" / "ingredientNames.js"


def slug(name: str) -> str:
    """Identifiant unique à partir du nom (minuscules, espaces → _, pas d'accents)."""
    s = (name or "").strip().lower()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[-\s]+", "_", s)
    return s[:80] or "ingredient"


def normalize_name(name: str) -> str:
    """Pour déduplication : minuscules, espaces normalisés."""
    return " ".join((name or "").strip().lower().split())


def load_existing() -> List[Dict[str, Any]]:
    if not INGREDIENTS_FILE.exists():
        return []
    with open(INGREDIENTS_FILE, encoding="utf-8") as f:
        return json.load(f)


def save_ingredients(items: List[Dict[str, Any]]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(INGREDIENTS_FILE, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    print(f"Écrit {len(items)} ingrédients dans {INGREDIENTS_FILE}")


def find_column(row: Dict[str, str], *candidates: str) -> Optional[str]:
    """Retourne la valeur de la première colonne existante (insensible à la casse)."""
    keys_lower = {k.strip().lower(): k for k in row.keys()}
    for c in candidates:
        c_lower = c.lower().strip()
        for k, orig in keys_lower.items():
            if c_lower in k or k in c_lower:
                return row.get(orig)
    return None


def parse_float(val: Any) -> float:
    if val is None or val == "" or val == "-":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


# Colonnes possibles Ciqual (noms variables selon export CSV et XLSX)
NAME_KEYS = ["alim_nom_fr", "aliment", "nom", "name", "food", "description"]
ENERGY_KEYS = ["energie (kcal)", "energie (kcal/100g)", "energy_kcal", "kcal", "energie", "calories"]
PROTEIN_KEYS = ["protéines", "proteines (g)", "protein_g", "proteins", "facteur jones"]
CARBS_KEYS = ["glucides (g)", "carbohydrates_g", "glucides", "glucide"]
FAT_KEYS = ["lipides (g)", "fat_g", "lipides", "lipide", "fat"]
FIBER_KEYS = ["fibres alimentaires", "fibres (g)", "fiber_g", "fibres", "fiber"]


def get_value_from_row(row: Dict[str, str], key_candidates: List[str]) -> Optional[str]:
    keys_lower = {k.strip().lower(): k for k in row.keys()}
    for c in key_candidates:
        c_lower = c.strip().lower()
        for k, orig in keys_lower.items():
            if c_lower in k or k in c_lower:
                return row.get(orig)
    return None


def row_to_per_100g(row: Dict[str, str]) -> Dict[str, float]:
    """Extrait energy_kcal, protein_g, carbohydrates_g, fat_g, fiber_g depuis une ligne CSV."""
    energy = parse_float(get_value_from_row(row, ENERGY_KEYS))
    # Ciqual peut fournir Energie en kJ uniquement
    if energy == 0:
        kj = get_value_from_row(row, ["energie (kj)", "energie (kj/100g)", "energy_kj"])
        if kj:
            energy = parse_float(kj) / 4.184  # kJ -> kcal
    protein = parse_float(get_value_from_row(row, PROTEIN_KEYS))
    carbs = parse_float(get_value_from_row(row, CARBS_KEYS))
    fat = parse_float(get_value_from_row(row, FAT_KEYS))
    fiber = parse_float(get_value_from_row(row, FIBER_KEYS))
    return {
        "energy_kcal": round(energy, 1),
        "protein_g": round(protein, 1),
        "carbohydrates_g": round(carbs, 1),
        "fat_g": round(fat, 1),
        "fiber_g": round(fiber, 1),
    }


def _rows_to_ingredients(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Convertit des lignes (dict) en liste d'ingrédients HealthTrack."""
    items = []
    seen_normalized = set()
    for row in rows:
        name_val = get_value_from_row(row, NAME_KEYS)
        if not name_val:
            continue
        name_val = " ".join(str(name_val).strip().split())  # collapse newlines/spaces
        if not name_val:
            continue
        norm = normalize_name(name_val)
        if norm in seen_normalized:
            continue
        seen_normalized.add(norm)
        per_100g = row_to_per_100g(row)
        items.append({
            "id": slug(name_val),
            "name": name_val,
            "per_100g": per_100g,
        })
    return items


def import_ciqual(csv_path: Path) -> List[Dict[str, Any]]:
    """Importe un CSV type Ciqual (séparateur ; ou ,, encodage UTF-8 ou Latin-1)."""
    for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(csv_path, encoding=encoding, newline="") as f:
                sample = f.read(4096)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"Impossible de décoder {csv_path}")

    delimiter = ";" if ";" in sample.split("\n")[0] else ","
    with open(csv_path, encoding=encoding, newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = list(reader)

    if not rows:
        print("Aucune ligne dans le CSV.")
        return []
    return _rows_to_ingredients(rows)


def import_ciqual_xlsx(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Importe la table Ciqual depuis un fichier Excel (.xlsx)."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Pour lire un fichier .xlsx, installez openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb.active
    if not sheet:
        wb.close()
        return []

    # Première ligne = en-têtes (normaliser: remplacer retours à la ligne par espace)
    header = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        v = (cell.value or "").strip()
        v = " ".join(v.split())
        header.append(v)

    rows = []
    for row in sheet.iter_rows(min_row=2):
        d = {}
        for i, cell in enumerate(row):
            if i < len(header) and header[i]:
                d[header[i]] = cell.value
        if d:
            rows.append(d)
    wb.close()

    if not rows:
        print("Aucune donnée dans la feuille Excel.")
        return []
    return _rows_to_ingredients(rows)


def merge_with_existing(
    existing: List[Dict[str, Any]],
    new_items: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Fusionne les nouveaux ingrédients sans dupliquer par nom normalisé."""
    by_norm = {normalize_name(i["name"]): i for i in existing}
    added = 0
    for it in new_items:
        norm = normalize_name(it["name"])
        if norm not in by_norm:
            by_norm[norm] = it
            added += 1
    result = list(by_norm.values())
    result.sort(key=lambda x: x["name"].lower())
    print(f"Ajout de {added} nouveaux ingrédients (total {len(result)}).")
    return result


def export_ingredient_names_js(items: List[Dict[str, Any]], max_names: int = 0) -> None:
    """Génère app/src/data/ingredientNames.js à partir de la base (max_names=0 = tous)."""
    names = [i["name"] for i in items]
    if max_names > 0:
        names = names[:max_names]
    content = (
        "/**\n"
        " * Noms canoniques d'ingrédients pour le prompt Gemini (standalone).\n"
        " * Généré par backend/scripts/enrich_ingredients.py --export-names\n"
        " * À partir de ingredients_nutrition.json (Ciqual + base existante).\n"
        " */\n"
        f"export const INGREDIENT_NAMES = [\n"
        + "  " + ",\n  ".join(repr(n) for n in names) + ",\n"
        "]\n"
    )
    APP_NAMES_FILE.parent.mkdir(parents=True, exist_ok=True)
    APP_NAMES_FILE.write_text(content, encoding="utf-8")
    print(f"Écrit {len(names)} noms dans {APP_NAMES_FILE}")


def main() -> None:
    if len(sys.argv) < 3:
        print(__doc__)
        print("Exemple: python -m scripts.enrich_ingredients ciqual path/to/Table_Ciqual.csv")
        sys.exit(1)

    cmd = (sys.argv[1] or "").strip().lower()
    path = Path(sys.argv[2])
    export_names = 0
    args = sys.argv[3:]
    for i, arg in enumerate(args):
        if arg == "--export-names":
            if i + 1 < len(args):
                try:
                    export_names = int(args[i + 1])
                except ValueError:
                    export_names = 500
            else:
                export_names = 500
            break

    if not path.exists():
        print(f"Fichier introuvable: {path}")
        sys.exit(1)

    existing = load_existing()

    if cmd == "ciqual":
        if path.suffix.lower() in (".xlsx", ".xls"):
            new_items = import_ciqual_xlsx(path)
        else:
            new_items = import_ciqual(path)
        print(f"Import Ciqual: {len(new_items)} lignes valides.")
        merged = merge_with_existing(existing, new_items)
        save_ingredients(merged)
        if export_names:
            export_ingredient_names_js(merged, export_names)
    elif cmd == "merge":
        # Merge un JSON au format HealthTrack (liste d'objets id, name, per_100g)
        with open(path, encoding="utf-8") as f:
            extra = json.load(f)
        if isinstance(extra, list):
            new_items = [e for e in extra if isinstance(e, dict) and e.get("name")]
        else:
            new_items = []
        merged = merge_with_existing(existing, new_items)
        save_ingredients(merged)
        if export_names:
            export_ingredient_names_js(merged, export_names)
    else:
        print("Commande inconnue. Utilisez: ciqual <fichier.csv> ou merge <fichier.json>")
        sys.exit(1)


if __name__ == "__main__":
    main()
